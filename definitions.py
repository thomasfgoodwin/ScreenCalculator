import math
from openpyxl import load_workbook
import math
from fractions import Fraction
resolutions = {}
res_table_headings = {}
index = {}
units = {
    'miles': 5280,
    "yards": 3,
    "feet": 1,
    'inches': 1 / 12,
}
unit_order = [ 1 / 12, 1 / 12]  # Read if no units provided. Default reads "feet, inches, inches"
aliases = {
    'foot': 'feet',
    'ft': 'feet',
    'inch': 'inches',
    'in': 'inches',
    '\'': 'feet',
    '\"': 'inches',
    'yd': 'yards',
    "yards": 'yards',
    'mi': 'miles',
    'miles': "miles"
}


def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False


def search_dictionary(q):
    query = q
    if query in index:
        i = index[query]
        w = (resolutions[i]['W'])
        h = (resolutions[i]['H'])
        return w, h
    else:
        print("QUERY NOT RECOGNIZED, USING DEFAULT (1920x1080)")
        return 1920, 1080


def build_dictionary(f, hints=True):
    worksheet = load_workbook(filename=f'{f}')
    sheet_ranges = worksheet['Data']
    x = sheet_ranges.max_column
    y = sheet_ranges.max_row
    last_cell = sheet_ranges['A1'].offset(y-1, x-1).coordinate
    cell = sheet_ranges[f'A1:{last_cell}']

    for row in range(0, len(cell)):
        if row == 0:
            for column in range(0, len(cell[0])):
                res_table_headings[column] = cell[row][column].value
        else:
            ID = cell[row][0].value
            string = str(cell[row][1].value)
            if string != 'None':
                aliases = string.split(', ')
                #tags.append(aliases)
                if hints:
                    print(aliases)
                for alias in aliases:
                    index[alias] = row
            index[ID] = row
            resolutions[row] = {}
            for column in range(0, len(cell[0])):
                attribute = res_table_headings[column]
                value = cell[row][column].value
                resolutions[row][attribute] = value


def arch_units(num, max_denominator=16):
    if max_denominator == 0:
        max_denominator = 1
    ft = math.floor(num)
    inch = math.floor((num - ft) * 12)
    error = (num - ft - (inch / 12)) * 12
    frac = round(error * max_denominator)
    error = (num - ft - (inch / 12) - (frac / (12 * max_denominator))) / (12 * max_denominator)
    if frac == max_denominator:
        inch = inch + 1
        frac = frac - max_denominator
    output = []
    ft_display = f"{ft}' "
    inch_display = f'{inch}'
    frac_display = f'{frac}/{max_denominator}'
    if ft != 0:
        output.append(ft_display)
    if inch != 0:
        output.append(inch_display)
    if inch != 0 and frac != 0:
        output.append('-')
    if frac != 0:
        output.append(frac_display)
    if frac + inch != 0:
        output.append('"')
    sep = ""
    output = str(sep.join(output))
    return output


def arch2dec(s):  ## input is any common format for imperial length measurement, ex: 6'4-/4", 10.5 mi, 8f oot 1
    s = s.strip()  # trim whitespace
    s = s.replace('-', ' ')
    indices = []
    n = 0
    mode = 'U'  # mode is U so first number increments n counter
    for i, c in enumerate(s):  # check every character
        number_bool = c.isnumeric() or c == '.'  # is it a number or '.'
        if mode == 'N':  # if we're reading a number
            if not number_bool:  # and we find a non-number
                mode = 'U'  # switch to reading a unit
                indices.append(i)  # and record the index of where this happened
        if mode == 'U':  # if we're reading a unit
            if number_bool:  # and we find a number
                mode = 'N'  # switch to reading a number
                n = n + 1  # increment the number of number-unit pairs we're dealing with
                indices.append(i)  # and record the index of where this happened
    indices.append(len(s))
    # print(i, s[i], mode)
    phrases = []
    #   Create phrases of entire number strings and strings of whatever is between them:
    for a, b in zip(indices, indices[1:]):
        phrase = s[a:b]
        phrase = phrase.strip()
        phrases.append(phrase)  # make a list of phrases
    #   If a non-number phrase is recognized as an alias for a unit, replace it with that unit
    phrases = [aliases.get(item, item) for item in phrases]  # lookup units
    phrases = [phrase for phrase in phrases if phrase != '']  # remove empty items
    # print(phrases, n)
    #   Find fractions by looking for '/ and evaluating it as division:
    for i in phrases:
        if i == '/':
            div_index = phrases.index('/')
            numerator = float(phrases[div_index - 1])
            denominator = float(phrases[div_index + 1])
            phrases[div_index - 1] = numerator / denominator
            del phrases[div_index]
            del phrases[div_index]
            n = n - 1
    #   Create separate lists for numbers and non-numbers:
    numbers = []
    factors = []
    total = 0
    #   Create and sum components
    for i, phrase in enumerate(phrases):
        d = (i, phrase)
        if isfloat(phrase):
            numbers.append(d)
        else:
            factors.append(d)
    # print(f'number, factors: {numbers}, {factors}. n = {n}')
    #   If list of units is empty, assume numbers are presented as feet>inches>inches
    #   example: 6 4 1/2 is interpreted as 6' 4-1/2"
    if not factors:
        # print('no units')
        for i, number in enumerate(numbers):
            total = total + unit_order[i] * float(number[1])
        return total
    #   Units affect numbers that are before themselves but after the previous unit,
    #   and numbers after themselves if they are the last unit
    step = 0
    multipy = 1
    for j, number in numbers:
        component = 0
        for index, factor in enumerate(factors):
            i, f = factor
            f = units[f]
            if step <= j < i:
                multipy = f
            # print(f'i={i} j={j}')
            # print(f'value is {number}, unit is {factor}, factor is,{f}')
            step = i
        component = component + float(number) * float(multipy)
        # print(f'component: {number} {factor[1]} * {multipy}')
        total = total + component
    return total


def resolution(s):
    build_dictionary('resolutions.xlsx', False)
    query = s
    w, h = search_dictionary(query)
    aspect = w / h
    return w, h, aspect


def width_and_height(d, aspect):
    tau = math.atan(1 / aspect)
    height = d * math.sin(tau)
    width = d * math.cos(tau)
    return float(width), float(height)


def pixel_size(w, h, res_w, res_h):
    # res_w, res_h, a = resolution()
    pix_w = w / res_w
    pix_h = h / res_h
    pix_m = max(pix_w, pix_h)
    return pix_w, pix_h


def view_distance(x):
    pix_m = x
    acuity = 1  # visual acuity in arcminutes
    alpha = math.radians(acuity / 60)
    distance = pix_m / (1 * math.tan(alpha))
    return float(distance)


def calc(d, s):
    messages = []
    d_input = d
    d = arch2dec(d_input)
    ####
    res_w, res_h, aspect = resolution(s)
    w, h = width_and_height(d, aspect)
    pix_w, pix_h = pixel_size(w, h, res_w, res_h)
    ####
    pix_m = max(pix_w, pix_h)
    distance = (view_distance(pix_m))
    aspect_ratio = Fraction(aspect).limit_denominator()
    aspect_ratio = str(aspect_ratio).replace("/",":")
    pix_m = arch_units(pix_m, 256)
    ####
    messages.append(f'Display is {(round(h, 2))} high by {(round(w, 2))} wide')
    messages.append(f'Resolution is {res_w} x {res_h} for a {round(aspect, 2)} or {aspect_ratio} aspect ratio')
    if round(pix_w, 3) == round(pix_h, 3):
        messages.append(f'Pixels are {pix_m} square')
    else:
        messages.append(f'Pixels are {pix_h} by {pix_w}')
    messages.append(f'Individual pixels are perceptible at a distance of {arch_units(distance)}')
    return messages
